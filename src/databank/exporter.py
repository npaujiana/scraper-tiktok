"""
DataBankExporter — Export data bank contents to Excel (.xlsx) files.

Uses openpyxl (already a project dependency) to create styled Excel workbooks.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich import print
from rich.text import Text

from .models import COLUMN_LABELS

if TYPE_CHECKING:
    from .database import DataBank

__all__ = ["DataBankExporter"]

# Style constants
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(
    start_color="2F5496", end_color="2F5496", fill_type="solid"
)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)

# All exportable tables
ALL_TABLES = [
    "contents", "comments", "users",
    "search_users", "search_lives", "hot_trends",
]

# Human-readable sheet names
SHEET_NAMES = {
    "contents": "Contents",
    "comments": "Comments",
    "users": "Users",
    "search_users": "Search Users",
    "search_lives": "Search Lives",
    "hot_trends": "Hot Trends",
}


class DataBankExporter:
    """Export data bank contents to Excel (.xlsx) files."""

    def __init__(self, databank: "DataBank"):
        self.databank = databank

    async def export_all(
        self,
        output_path: Path | str,
    ) -> Path | None:
        """Export all data bank tables to a single multi-sheet Excel file.

        Args:
            output_path: Path for the output .xlsx file.

        Returns:
            The output path if successful, None otherwise.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            total_rows = 0
            for table in ALL_TABLES:
                data = await self.databank.get_all_data(table)
                if data:
                    sheet_name = SHEET_NAMES.get(table, table)
                    self._write_sheet(wb, sheet_name, table, data)
                    total_rows += len(data)

            if total_rows == 0:
                print(
                    Text(
                        "Data Bank is empty, nothing to export",
                        style="bold yellow",
                    )
                )
                return None

            wb.save(str(output_path))
            print(
                Text(
                    f"Exported {total_rows} records to {output_path}",
                    style="bold green",
                )
            )
            return output_path

        except Exception as e:
            print(
                Text(
                    f"Export failed: {e}",
                    style="bold red",
                )
            )
            return None

    async def export_by_type(
        self,
        data_type: str,
        output_path: Path | str,
        filters: dict | None = None,
    ) -> Path | None:
        """Export a specific data type to an Excel file.

        Args:
            data_type: Table name (contents, comments, users, etc.)
            output_path: Path for the output .xlsx file.
            filters: Optional dict of column -> value for filtering.

        Returns:
            The output path if successful, None otherwise.
        """
        if data_type not in ALL_TABLES:
            print(
                Text(
                    f"Unknown data type: {data_type}",
                    style="bold red",
                )
            )
            return None

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            data = await self.databank.get_all_data(data_type, filters)

            if not data:
                print(
                    Text(
                        f"No data found for type '{data_type}'",
                        style="bold yellow",
                    )
                )
                return None

            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAMES.get(data_type, data_type)
            self._write_sheet(wb, ws.title, data_type, data)
            # Remove extra default sheet if created
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            wb.save(str(output_path))
            print(
                Text(
                    f"Exported {len(data)} {data_type} records to {output_path}",
                    style="bold green",
                )
            )
            return output_path

        except Exception as e:
            print(
                Text(
                    f"Export failed: {e}",
                    style="bold red",
                )
            )
            return None

    async def export_filtered(
        self,
        output_path: Path | str,
        date_from: str | None = None,
        date_to: str | None = None,
        nickname: str | None = None,
        source_type: str | None = None,
        table: str = "contents",
    ) -> Path | None:
        """Export data with advanced filters.

        Args:
            output_path: Path for the output .xlsx file.
            date_from: Start date string (collection_time >=).
            date_to: End date string (collection_time <=).
            nickname: Filter by nickname.
            source_type: Filter by source_type.
            table: Table to query.

        Returns:
            The output path if successful, None otherwise.
        """
        if not self.databank.is_available:
            return None

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            # Build filtered query
            where_parts = []
            params = []
            idx = 1

            if date_from:
                where_parts.append(f"collection_time >= ${idx}")
                params.append(date_from)
                idx += 1

            if date_to:
                where_parts.append(f"collection_time <= ${idx}")
                params.append(date_to)
                idx += 1

            if nickname:
                where_parts.append(f"nickname ILIKE ${idx}")
                params.append(f"%{nickname}%")
                idx += 1

            if source_type:
                where_parts.append(f"source_type = ${idx}")
                params.append(source_type)
                idx += 1

            where_clause = (
                "WHERE " + " AND ".join(where_parts) if where_parts else ""
            )

            sql = (
                f"SELECT * FROM {table} {where_clause} "
                f"ORDER BY created_at DESC"
            )

            async with self.databank.pool.acquire() as conn:
                rows = await conn.fetch(sql, *params)
                data = [dict(row) for row in rows]

            if not data:
                print(
                    Text(
                        "No data matches the given filters",
                        style="bold yellow",
                    )
                )
                return None

            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAMES.get(table, table)
            self._write_sheet(wb, ws.title, table, data)
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            wb.save(str(output_path))
            print(
                Text(
                    f"Exported {len(data)} filtered records to {output_path}",
                    style="bold green",
                )
            )
            return output_path

        except Exception as e:
            print(
                Text(
                    f"Filtered export failed: {e}",
                    style="bold red",
                )
            )
            return None

    # ========================================================================
    # Private helper methods
    # ========================================================================

    def _write_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        table: str,
        data: list[dict],
    ):
        """Write data to a named sheet with styling."""
        ws = wb.create_sheet(title=sheet_name)

        if not data:
            return

        # Get column names from the first row
        columns = list(data[0].keys())

        # Skip internal columns (extra_data JSONB)
        skip_cols = {"extra_data"}
        columns = [c for c in columns if c not in skip_cols]

        # Write header row with labels
        labels = COLUMN_LABELS.get(table, {})
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(
                row=1,
                column=col_idx,
                value=labels.get(col_name, col_name),
            )
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                # Convert datetime/special types to string
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                elif isinstance(value, (list, dict)):
                    value = str(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT

        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            header_len = len(labels.get(col_name, col_name))
            max_len = header_len

            # Sample first 50 rows for width calculation
            for row_idx in range(2, min(len(data) + 2, 52)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cell_len = len(str(cell_value))
                    max_len = max(max_len, min(cell_len, 60))

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_len + 3

        # Freeze top row
        ws.freeze_panes = "A2"

    @staticmethod
    def generate_filename(prefix: str = "databank_export") -> str:
        """Generate a timestamped filename for exports."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"
